from PDF_to_Excel_Functions import * #External python program containing functions
import openpyxl #Used for performing excel operations

# The green hashtag text are comments for documenting purposes. They do not impact the code whatsoever. There will be many comments but you can delete them if they flood too much of the screen.
"""
The triple quotes are also comments. But I will mainly use this to document seperate sections of the code.
"""

#Yellow functions that start with a capital such as ExtractAllPDFs are custom functions I made in PDF_to_Excel_Functions.py; Which is an external python file.

"""
These are the base variables that are initialized first. They are the external files that are used.
"""
#This is my own file path. Will be different on different devices. Change the paths to accomodate for your own device.
path = "C:\\Users\\DyJo\\OneDrive - Government of Ontario\\Documents\\Visual Studio Code\\Python\\I&A PDF to Excel" #Base Path
excelFile = path + "\\Portal Applications Coded (1) (For Coding Purposes).xlsx" #Path of the excel file.
pdfPath = path + "\\Touch_Point_re_Scraping_PDFs" #Path where every PDF is located
workBook = openpyxl.load_workbook(excelFile) #workBook is the variable that holds the excel file.
sheet = workBook.active #sheet is the current selected sheet in the Excel Workbook(workBook). It is important that 'Main' is selected
pdfList = ExtractAllPDFs(pdfPath) #pdfList is set to function ExtractAllPDFs. Said function returns a list.

"""
This loop is infinite as the user may want to use various instructions as many times as they want.
"""
while True: #START WHILE TRUE LOOP
    userInput = str(input("Enter a value to perform an instruction: \
        \n0 for PDF to Excel \
        \n1 for populating form data \
        \n2 for difference days (B and N)\
        \n3 for diference days (B and L)\
        \n4 for year month format\
        \n5 for counting frequency of month/year combination\
        \n6 for counting frequency of Requestor Type\
        \nT to terminate\n")) #User is prompted to enter a value to determine which operation to perform
    if userInput == "0": #0 is the population of all of the available excel files in the folder designated by pdfPath
        for c in range(len(pdfList)): #Loops through the list for every PDF
            inputPDF = pdfList[c] #inputPDF is set to the current PDF being written into excel.
            titlePDF = inputPDF.split("_") #The title is split by the underscore to retrieve the portal number and date
            portalNumber = str(titlePDF[0]) #First half is the portal number
            date = str(FormatDate(titlePDF[1])) #Second half is the date.
            inputPDFPath = path + "\\Touch_Point_re_Scraping_PDFs\\" + inputPDF #Path of a specific PDF. Not to be confused with pdfPath
            textFile = ConvertPDFtoText(inputPDFPath, inputPDF, path) #Makes text file from the PDF.
            row = CurrentBlankRow(sheet, 4, 1, 50) #Retrieves the current blank row that can be populated
            ConvertTexttoExcel(textFile, sheet, row, portalNumber, date) #Writes the text from the text file into excel
            workBook.save(excelFile) #The excel file is saved so that it updates after a form is fully populated into the sheet.
            print(inputPDF + " written. (" + str(c + 1) + "/" + str(len(pdfList)) + ")") #Output line to indicate the PDF has been written to excel. Not really necessary but useful to determine progress of script.
        workBook.save(excelFile) #Save workbook one more time before ending operation.
        askUser = str(input("Do you want to keep the text files? Type 'Yes' to keep. Type anything else to delete:\n")) #Asks user whether to keep text files or not
        if askUser.lower() == "yes":
            print("Text Files Not Deleted.")
        else:
            DeleteTextFiles(pdfPath) #Function called to delete all text files.
            print("Text Files Deleted.")
        print("Done")
    elif userInput == "1": #1 is to populate the form data from an external excel document.
        excelFileTwo = path + "\\CDR Portal Survey Coding (1).xlsx" #This excel file is the one containing the survey results.
        workBookTwo = openpyxl.load_workbook(excelFileTwo)
        sheetTwo = workBookTwo.active
        currentRow1 = CurrentBlankRow(sheet, 4, 1, 50) #Retrieve current blank row of main excel document
        r1 = 4 #Row 4 contains the first populated PDF document.
        while r1 < currentRow1: #Loop through all PDF documents
            r2 = 17 #Row 17 contains the first PDF survey result
            currentRow2 = CurrentBlankRow(sheet, 17, 2, 14) #Retrieve current blank row of the survey result excel document
            while r2 < currentRow2: #Loop through all surey results. Nested loop
                CompareCells(sheet, sheetTwo, r1, r2) #Function to compare current survey result with current pdf from main document in outer loop
                workBook.save(excelFile) #Save main excel file, whether changes were made or not
                r2 += 1 #Increment 1 for next survey form
            print(str(r1) + "/" + str(currentRow1 - 1) + " completed.") #Not necessary but used to indicate progress
            r1 += 1 #Increment 1 for next pdf file
        workBookTwo.save(excelFileTwo) #Save survey result excel file to ensure safety
    elif userInput == "2": #2 is for calculating the difference in values between columns B and N
        currentRow = CurrentBlankRow(sheet, 4, 1, 50)
        r = 4
        while r < currentRow:
            DifferenceDays(sheet, r, 70, 2, 14) #Performs calculation
            workBook.save(excelFile)
            print("Row " + str(r) + " Updated")
            r += 1
    elif userInput == "3": #3 is for calculating the difference in values between columns B and L
        currentRow = CurrentBlankRow(sheet, 4, 1, 50)
        r = 4
        while r < currentRow:
            DifferenceDays(sheet, r, 71, 2, 12)
            workBook.save(excelFile)
            print("Row " + str(r) + " Updated")
            r += 1
    elif userInput == "4": #4 is for formatting the years into MM/YYYY format
        currentRow = CurrentBlankRow(sheet, 4, 1, 50)
        r = 4
        while r < currentRow:
            FormatMonthYear(sheet, r) #Performs formatting
            workBook.save(excelFile)
            print("Row " + str(r) + " Updated")
            r += 1
    elif userInput == "5": #5 is for counting the frequency of the month/year combinations for a specified requestor type. Only reads the excel file and does not edit it
        requestorType = str(input("Enter Requestor Type. Type 'All' for all requestor types:\n"))
        currentRow = CurrentBlankRow(sheet, 4, 1, 50)
        r = 4
        print(str(CountCombinations(sheet, r, currentRow, 72, requestorType)))
        workBook.save(excelFile)
    elif userInput == "6": #6 is for counting the number of forms for each requestor type. Just like 6, it does not edit the excel file but only reads it
        currentRow = CurrentBlankRow(sheet, 4, 1, 50)
        r = 4
        print(str(CountCombinations(sheet, r, currentRow, 18, "All")))
        workBook.save(excelFile)
    elif userInput == "T": #T is for when you want to terminate (end) the program. It will break out of the infinite loop
        break
    else: #Any other form of input is not accepted and will be declared as bad input. Will be prompt to enter a valid input
        print("\nBad Input\n")
    workBook.save(excelFile) #Another save to ensure the excel file's safety
#END WHILE TRUE LOOP
workBook.save(excelFile) #One final save to ensure the excel file's safety
print("Program Terminated") #Used to indicate that the program has ended since this is the last line in the program.